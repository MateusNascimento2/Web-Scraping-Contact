from __future__ import barry_as_FLUFL
from this import d
import time
from tkinter import filedialog, ttk
from turtle import width
from numpy import save
from openpyxl import load_workbook
import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from pyparsing import col
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from tkinter import filedialog as fd
from tkinter import *
from tkinter.messagebox import showinfo
import tkinter as tk
from tkinter import ttk, Canvas, NW
from tkinter.ttk import Progressbar
from tkinter.messagebox import showerror
from PIL import ImageTk, Image
import datetime



win = Tk()



win.title("")
win.geometry("1280x720")
ico = Image.open("img/icone-logo.png")
photo_icon = ImageTk.PhotoImage(ico)
win.wm_iconphoto(False, photo_icon)
#win.configure(bg='#040405')
win.resizable(0, 0)


photo = Image.open("img/background.jpg")
resized_image = photo.resize((1600,900), Image.Resampling.LANCZOS)
new_image= ImageTk.PhotoImage(resized_image)
width, height = new_image.width(), new_image.height()
canvas = Canvas(win, width=width, height=height, highlightthickness=0)
canvas.pack()
canvas.create_image(0, 0, image=new_image, anchor=NW)

photo2 = Image.open("img/logo-3.png")
resized_image2 = photo2.resize((270, 75), Image.Resampling.LANCZOS)
new_image2 = ImageTk.PhotoImage(resized_image2)
width, height = new_image2.width(), new_image2.height()
canvas2 = Canvas(win, bg="white", width=width, height=height, highlightthickness=0)
canvas2.pack()
canvas2.create_image(0, 0, image=new_image2, anchor=NW)
canvas2.place(x=500, y=10)


## --------------LEMIT GUI------------------------ ##

credencias_label = Label(win, text="Credenciais", bg="white", fg='black', font=('yu gothic ui', 17, 'bold'))
credencias_label.place(x=150, y=120)

sign_in_label = Label(win, text='Lemit', bg="white", fg='black', font=('yu gothic ui', 15, 'bold'))
sign_in_label.place(x=150, y=170)

email_label = Label(win, text="Email", bg='white', fg='black', font=('yu gothic ui', 14, 'bold'))
email_label.place(x=150, y=225)

email_entry = Entry(win, highlightthickness=0, relief=FLAT, bg="#e6e9ed", width=25, font=('yu gothic ui', 13))
email_entry.place(x=155, y=255)

senha_label = Label(win, text="Senha", bg="white", fg='black', font=('yu gothic ui', 14, 'bold'))
senha_label.place(x=150, y=295)

senha_entry = Entry(win, highlightthickness=0, relief=FLAT, bg="#e6e9ed", width=25, show="*", font=('yu gothic ui', 13, 'bold'))
senha_entry.place(x=155, y=325)


## --------------Assertiva GUI------------------------ ##

sign_in_label = Label(win, text='Assertiva', bg="white", fg='black', font=('yu gothic ui', 15, 'bold'))
sign_in_label.place(x=497, y=170)

email_label = Label(win, text="Email", bg='white', fg='black', font=('yu gothic ui', 14, 'bold'))
email_label.place(x=496, y=225)

email_entry2 = Entry(win, highlightthickness=0, relief=FLAT, bg="#e6e9ed", width=25, font=('yu gothic ui', 13))
email_entry2.place(x=500, y=255)

senha_label = Label(win, text="Senha", bg="white", fg='black', font=('yu gothic ui', 14, 'bold'))
senha_label.place(x=496, y=295)

senha_entry2 = Entry(win, highlightthickness=0, relief=FLAT, bg="#e6e9ed", width=25, show="*", font=('yu gothic ui', 13, 'bold'))
senha_entry2.place(x=500, y=325)

## --------------Confirme Online GUI------------------------ ##

sign_in_label = Label(win, text='Confirme Online', bg="white", fg='black', font=('yu gothic ui', 15, 'bold'))
sign_in_label.place(x=856, y=170)

email_label = Label(win, text="Email", bg='white', fg='black', font=('yu gothic ui', 14, 'bold'))
email_label.place(x=855, y=225)

email_entry3 = Entry(win, highlightthickness=0, relief=FLAT, bg="#e6e9ed", width=25, font=('yu gothic ui', 13))
email_entry3.place(x=860, y=255)

senha_label = Label(win, text="Senha", bg="white", fg='black', font=('yu gothic ui', 14, 'bold'))
senha_label.place(x=855, y=295)

senha_entry3 = Entry(win, highlightthickness=0, relief=FLAT, bg="#e6e9ed", width=25, show="*", font=('yu gothic ui', 13, 'bold'))
senha_entry3.place(x=860, y=325)

## botoes, barra de progresso ## 
excel_file_label = ttk.Label(win, text="Arquivo: ", font=(('yu gothic ui', 12, 'bold')), background='white')
excel_file_label.place(x=155, y=520)

label_file = ttk.Label(win, text="Nenhum arquivo selecionado", font=(('yu gothic ui', 12, 'bold')), background='white')
label_file.place(x=225, y=520)

button_escolher = tk.Button(win, text='Escolher Planilha', bd=0, fg='black', font=('yu gothic ui', 11, 'bold'), cursor='hand2', bg='#f5f8fa', activebackground='black', command=lambda: File_dialog())
button_escolher.place(x=155, y=550)

button_enviar = tk.Button(win, text="Enviar Planilha", bg='#f5f8fa', fg='black', font=('yu gothic ui', 11, 'bold'), borderwidth=1, cursor='hand2', relief='flat', activebackground='black', command=lambda: Load_excel_data())
button_enviar.place(x=300, y=550)

label_file_save = ttk.Label(win, text="Nenhum local selecionado", font=(('yu gothic ui', 12, 'bold')), background='white')
label_file_save.place(x=205, y=400)

button_salvar = tk.Button(win, text="Escolha um local para salvar o arquivo", bg='#f5f8fa', fg='black', font=('yu gothic ui', 11, 'bold'), borderwidth=1, cursor='hand2', relief='flat', activebackground='black', command=lambda: save())
button_salvar.place(x=150, y=440)

label_local = ttk.Label(win, text="Local:", font=(('yu gothic ui', 12, 'bold')), background='white')
label_local.place(x=155, y=400)


#bar = Progressbar(win, orient=HORIZONTAL, length=1280)
#bar.place(x=0, y=698)


def File_dialog():
    filename = filedialog.askopenfilename(initialdir="/", title="Select A File", filetype=(("xlsx files", "*.xlsx"),("All files", "*.*")))
    label_file['text'] = filename

def get_login():
    login = email_entry.get()
    return login

def get_senha():
    senha = senha_entry.get()
    return senha


def save():
    now = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
    global save_spot
    save_spot = fd.askdirectory()
    save_spot = str(save_spot)+f'/Planilha Atualizada {now}.xlsx'
    label_file_save['text'] = save_spot

    

def Load_excel_data():
    file_path = label_file['text']
    try:
        excel_filename = r"{}".format(file_path)
        data_excel = pd.read_excel(excel_filename,usecols=['Unnamed: 4', "Unnamed: 6"], skiprows=2).dropna()
        data_excel.drop(data_excel[data_excel["Unnamed: 6"] == "Escritório"].index, inplace=True )
        data_excel.drop(data_excel[data_excel["Unnamed: 6"] == "Escritorio"].index, inplace=True)
        data_excel.drop(data_excel[data_excel["Unnamed: 6"] == "Empresa"].index, inplace=True)
        data_excel.drop(data_excel[data_excel['Unnamed: 4'] == "CANCELADO"].index, inplace=True)
        data_excel.astype({"Unnamed: 4": float})
        df = data_excel.apply(lambda row: row[data_excel["Unnamed: 4"] >= 50000])
        
        cpfs = df["Unnamed: 6"].tolist()
        cpfs = list(dict.fromkeys(cpfs))
        for cpf in cpfs:
            if cpf == "Empresa":
                cpfs.remove(cpf)
            if cpf == "Escritorio":
                cpfs.remove(cpf)
            if cpf == "Escritório":
                cpfs.remove(cpf)
        
        df["Unnamed: 6"] = df["Unnamed: 6"].astype('str').str.replace(".", "")
        df["Unnamed: 6"] = df["Unnamed: 6"].astype('str').str.replace("-", "")
        index_row = df["Unnamed: 6"].index.tolist()
        
        new_index_row = [x+4 for x in index_row]
        new_index_row_str = [str(x) for x in new_index_row]
        
        cpf_list = df["Unnamed: 6"].tolist()
        
        for cpf in cpf_list:
            if cpf == "Empresa":
                cpf_list.remove(cpf)
            if cpf == "Escritorio":
                cpf_list.remove(cpf)
            if cpf == "Escritório":
                cpf_list.remove(cpf)
            if cpf == "CANCELADO":
                cpf_list.remove(cpf)
        print(len(cpf_list))
        print(len(new_index_row_str))
        
        url = "https://lemitti.com/auth/login"
        
        #chrome_options = Options()
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
        
        
        driver.get(url)
        
        
        driver.implicitly_wait(5)
        try:
            driver.find_element("name", "email").send_keys(get_login())
            driver.find_element("name", "password").send_keys(get_senha())
            driver.find_element("xpath", "/html/body/fieldset/div/div/div/form/div/div/div[4]/button").click()
            driver.find_element("xpath", "/html/body/fieldset/nav/div/div[2]/ul/li[3]/a").click()
            driver.find_element("xpath", "/html/body/fieldset/nav/div/div[2]/ul/li[3]/ul/li[1]/a").click()
        except:
            tk.messagebox.showerror(title="Erro", message="Email ou Senha Inválidos")
            return None
        
        planilha = load_workbook(file_path)
        aba_ativa = planilha.active
        
        contador = 0
        for (cpf, value) in zip(cpf_list, new_index_row_str):
            print(value, cpf)
            label_count = ttk.Label(win, text=str(len(cpf_list)), font=(('yu gothic ui', 12, 'bold')), background='white')
            label_count.place(x=185, y= 650)
            label_barra = ttk.Label(win, text=" |",  font=(('yu gothic ui', 12, 'bold')), background='white')
            label_barra.place(x=165, y= 650)
            contador += 1
            label_count2 = ttk.Label(win, text=str(contador), font=(('yu gothic ui', 12, 'bold')), background='white')
            label_count2.place(x=145, y=650)
            win.update()
            try:
                print(len(cpf_list))
                driver.get("https://lemitti.com/queries/cpf/" + cpf)
                time.sleep(10)
                driver.find_element("xpath", "/html/body/fieldset/div[2]/div[4]/div/div/div/form/div/div/span/button").click()
                element_nome = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[3]/div[2]/div[1]/table")
               
                element_celular = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[6]/div[2]/div/table")
                element_fixo = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[7]/div[2]/div/table")

                html_content_nome = element_nome.get_attribute('outerHTML')
                html_content_celular = element_celular.get_attribute('outerHTML')
                html_content_fixo = element_fixo.get_attribute('outerHTML')
                soup_nome = BeautifulSoup(html_content_nome, 'html.parser')
                soup_celular = BeautifulSoup(html_content_celular, 'html.parser')
                soup_fixo = BeautifulSoup(html_content_fixo, 'html.parser')
                table1 = soup_nome.find(name='table')
                table = soup_celular.find(name='table')
                table2 = soup_fixo.find(name='table')
                # print(table)
                # print(table2)
                df_full1 = pd.read_html(str(table1))[0]
                df_full = pd.read_html(str(table))[0]
                df_full2 = pd.read_html(str(table2))[0]
                nome = df_full1[1][0]
            
            

                lista_numero_celular = df_full["Número"].values.tolist()
                lista_numero_fixo = df_full2["Número"].values.tolist()
                new_lista_numero_celular = []
                for n in lista_numero_celular:
                    new_lista_numero_celular.append(n)
                new_lista_numero_fixo = []
                for n1 in lista_numero_fixo:
                    new_lista_numero_fixo.append(n1)
                lista_numero = [*new_lista_numero_celular, *new_lista_numero_fixo]
                print(nome)
                #print(new_lista_numero_celular)
                #print(new_lista_numero_fixo)
                print(lista_numero)

                link_table = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[4]/div[2]/div/table")
                html_content = link_table.get_attribute('outerHTML')
                soup = BeautifulSoup(html_content, 'html.parser')
                cpf_table = soup.find(name='table')
                df_full3 = pd.read_html(str(cpf_table))[0]
                nome_vinculo_list = df_full3["Nome"].values.tolist()
                tipo_vinculo_list = df_full3["Tipo de vínculo"].values.tolist()
                new_lista_vinculo = []
                i = 0
                for value1 in nome_vinculo_list:
                    while i < len(tipo_vinculo_list):
                        new_lista_vinculo.append(value1 + ' (' + tipo_vinculo_list[i] + ')')
                        i += 1
                        break
                print(new_lista_vinculo)
                df_full3["CPF"] = df_full3["CPF"].astype('str').str.replace('.', "", regex=True)
                df_full3["CPF"] = df_full3["CPF"].astype('str').str.replace('-', "", regex=True)
                cpf_vinculo_list = df_full3['CPF'].values.tolist()
                for v in cpf_vinculo_list:
                    if v == "nenhum vínculo encontrado":
                        cpf_vinculo_list.remove(v)
                #print(len(cpf_vinculo_list))
            
                #new_lista_numero_celular2 = []
                #new_lista_numero_fixo2 = []
                lista_numero2 = []
            
                if len(cpf_vinculo_list) == 0:
                    lista_numero2 = []
                    lista_numero2.append("nenhum telefone encontrado")
                    print(lista_numero2)

                if len(cpf_vinculo_list) <= 3:
                    v = 0
                    for x in cpf_vinculo_list:
                        driver.get("https://lemitti.com/queries/cpf/" + x)
                        time.sleep(10)
                        try:
                            driver.find_element("xpath", "/html/body/fieldset/div[2]/div[4]/div/div/div/form/div/div/span/button").click()
                            element_nome2 = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[3]/div[2]/div[1]/table")
                            element_celular2 = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[6]/div[2]/div/table")

                            element_fixo2 = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[7]/div[2]/div/table")
                            html_content_nome2 = element_nome2.get_attribute('outerHTML')

                            html_content_celular2 = element_celular2.get_attribute('outerHTML')
                            html_content_fixo2 = element_fixo2.get_attribute('outerHTML')
                            soup_nome2 = BeautifulSoup(html_content_nome2, 'html.parser')
                            soup_celular2 = BeautifulSoup(html_content_celular2, 'html.parser')
                            soup_fixo2 = BeautifulSoup(html_content_fixo2, 'html.parser')
                            table3 = soup_celular2.find(name='table')
                            table4 = soup_fixo2.find(name='table')
                            table5 = soup_nome2.find(name='table')

                            df_full4 = pd.read_html(str(table3))[0]
                            df_full5 = pd.read_html(str(table4))[0]
                            df_full6 = pd.read_html(str(table5))[0]
                            nome2 = df_full6[1][0]
                            nome2_vinculo = nome2 + " (" + tipo_vinculo_list[v] + ")"

                            lista_numero_celular2 = df_full4["Número"].values.tolist()
                            lista_numero_fixo2 = df_full5["Número"].values.tolist()
                    
                            lista_numero2.append(nome2_vinculo)
                            for n in lista_numero_celular2:
                                lista_numero2.append(n)
                            for n2 in lista_numero_fixo2:
                                lista_numero2.append(n2)
                            #print(new_lista_numero_celular2)
                            #print(new_lista_numero_fixo2)
                            #lista_numero2 = [*new_lista_numero_celular2, *new_lista_numero_fixo2]
                            v += 1 
                            print(lista_numero2)
                        except:
                            if len(lista_numero2) == 0:
                                lista_numero2 = []
                                lista_numero2.append("cpf do vínculo não encontrado")
                                v += 1 
                            else:
                                lista_numero2.append("cpf do vínculo não encontrado")
                                v += 1
                        continue
                    

                if len(cpf_vinculo_list) > 3:
                    i = 0
                    v = 0
                    while i <= 2:
                        driver.get("https://lemitti.com/queries/cpf/" + cpf_vinculo_list[i])
                        time.sleep(10)
                        try:
                            driver.find_element("xpath", "/html/body/fieldset/div[2]/div[4]/div/div/div/form/div/div/span/button").click()
                            element_nome2 = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[3]/div[2]/div[1]/table")
                            element_celular2 = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[6]/div[2]/div/table")
                            element_fixo2 = driver.find_element("xpath", "/html/body/fieldset/div[2]/div[3]/div/div[7]/div[2]/div/table")
                            html_content_nome2 = element_nome2.get_attribute('outerHTML')
                            html_content_celular2 = element_celular2.get_attribute('outerHTML')
                            html_content_fixo2 = element_fixo2.get_attribute('outerHTML')
                            soup_nome2 = BeautifulSoup(html_content_nome2, 'html.parser')
                            soup_celular2 = BeautifulSoup(html_content_celular2, 'html.parser')
                            soup_fixo2 = BeautifulSoup(html_content_fixo2, 'html.parser')
                            table3 = soup_celular2.find(name='table')
                            table4 = soup_fixo2.find(name='table')
                            table5 = soup_nome2.find(name='table')

                            df_full4 = pd.read_html(str(table3))[0]
                            df_full5 = pd.read_html(str(table4))[0]
                            df_full6 = pd.read_html(str(table5))[0]
                            nome2 = df_full6[1][0]
                            nome2_vinculo = nome2 + " (" + tipo_vinculo_list[v] + ")"

                            lista_numero_celular2 = df_full4["Número"].values.tolist()
                            lista_numero_fixo2 = df_full5["Número"].values.tolist()
                            print(lista_numero_celular2)
                    
                            lista_numero2.append(nome2_vinculo)
                            for n in lista_numero_celular2:
                                lista_numero2.append(n)
                            for n2 in lista_numero_fixo2:
                                lista_numero2.append(n2)
                        
                            #print(new_lista_numero_celular2)
                            #print(new_lista_numero_fixo2)
                            #lista_numero2 = [*new_lista_numero_celular2, *new_lista_numero_fixo2]
                            print(lista_numero2)
                            v += 1
                            i += 1
                        except:
                            if len(lista_numero2) == 0:
                                lista_numero2.append("cpf do vínculo não encontrado")
                                i += 1
                                v += 1
                            else:
                                lista_numero2.append("cpf do vínculo não encontrado")
                                i += 1
                                v += 1
                        continue
                        
            
            

                if len(new_lista_vinculo) >= 4:
                    today = datetime.datetime.now().strftime('%d-%m-%Y')
                    lista_numero_all = [*lista_numero, lista_numero2]
                    lista_numero_all = [nome] + lista_numero_all + [f"Atualizado em: {today}"]
                    print(lista_numero_all)
                    
                    aba_ativa["H" + value] = repr(lista_numero_all)
            
                if len(new_lista_vinculo) == 3: 
                    new_lista_vinculo3 = new_lista_vinculo[0:3]
                    #print(new_lista_vinculo3)
                    #print(new_lista_vinculo3[0])
                    today = datetime.datetime.now().strftime('%d-%m-%Y')
                    lista_numero_all = [*lista_numero, lista_numero2]
                    lista_numero_all = [nome] + lista_numero_all + [f"Atualizado em: {today}"]
                    print(lista_numero_all)
           

                    aba_ativa["H" + value] = repr(lista_numero_all)
            
            

                if len(new_lista_vinculo) == 2: 
                    new_lista_vinculo3 = new_lista_vinculo[0:2]
                    #print(new_lista_vinculo3)
                    #print(new_lista_vinculo3[0])
                    today = datetime.datetime.now().strftime('%d-%m-%Y')
                    lista_numero_all = [*lista_numero, lista_numero2]
                    lista_numero_all = [nome] + lista_numero_all + [f"Atualizado em: {today} "]
                    print(lista_numero_all)

                    aba_ativa["H" + value] = repr(lista_numero_all)
                     
                
                if len(new_lista_vinculo) == 1:
                    new_lista_vinculo3 = new_lista_vinculo[0:1]
                    #print(new_lista_vinculo3)
                    #print(new_lista_vinculo3[0])
                    today = datetime.datetime.now().strftime('%d-%m-%Y')
                    lista_numero_all = [*lista_numero, lista_numero2]
                    lista_numero_all = [nome] + lista_numero_all + [f"Atualizado em: {today} "]
                    print(lista_numero_all)
                
                    aba_ativa["H" + value] = repr(lista_numero_all)
                
                    
            except:
                print("cpf nao encontrado")
                today = datetime.datetime.now().strftime('%d-%m-%Y')
                aba_ativa["H" + value] = "cpf não encontrado. " +  f"Atualizado em: {today}" 
            continue
    
        planilha.save(save_spot)
        
        
        tk.messagebox.showinfo(message="Operação concluída com sucesso")
        driver.quit()
        
    except ValueError:
        tk.messagebox.showerror(title="Erro", message="Arquivo inválido")
        return None
    except FileNotFoundError:
        tk.messagebox.showerror(title="Erro", message="Nenhum arquivo encontrado")
        return None
    

win.mainloop()



